import sqlite3
import csv
import openpyxl
import math


class NoSelectedTable(Exception):
    pass


class NoSuchColumns(Exception):
    pass


class NoSuchTable(Exception):
    pass


class Table:
    def __init__(self, filename, *columns, format=''):
        self.columns = columns
        self.filename = filename
        self.table = []
        if format == '.sqlite':
            self.sql_table()
        elif format == '.csv':
            self.csv_table()
        elif format == '.xlsx':
            self.excel_table()
        else:
            raise NoSelectedTable

    def sql_table(self):
        table_name = self.columns[0]
        self.columns = self.columns[1:]
        if self.columns:
            response = f'''SELECT {', '.join(self.columns)} from {table_name}'''
            con = sqlite3.connect(self.filename)
            cur = con.cursor()
            table = list(cur.execute(response))
            for line in table:
                for element in line:
                    self.table += [element]
        else:
            response = f'SELECT * from {table_name}'
            con = sqlite3.connect(self.filename)
            cur = con.cursor()
            table = list(cur.execute(response))
            for line in table:
                for element in line:
                    self.table += [element]

    def csv_table(self):
        with open(self.filename, encoding="utf8") as file:
            reader = csv.reader(file, delimiter=';', quotechar='"')
            if self.columns:
                for line in reader:
                    for column in self.columns:
                        self.table.append(line[column])
            else:
                for line in reader:
                    self.table += line

    def excel_table(self):
        workbook = openpyxl.load_workbook(self.filename)
        worksheet = workbook.active
        all_data = []
        for line in worksheet:
            for obj in line:
                all_data.append(obj.value)
        table_data = []
        for i in range(worksheet.max_row):
            table_data += [all_data[:worksheet.max_column]]
            all_data = all_data[worksheet.max_column:]
        table = list(zip(*table_data))
        if self.columns:
            selected_columns = list(filter(lambda k: table.index(k) in self.columns, table))
            for column in selected_columns:
                self.table += list(column)

        else:
            for line in table:
                for element in line:
                    self.table += [element]

    def get_table(self):
        return self.table


class StatisticalAnalysis:
    def __init__(self, table):
        self.table = [int(element) for element in table if str(element).isdigit()]
        self.len = len(self.table)

    def length(self):
        return self.len

    def amount(self):
        return sum(self.table)

    def arithmetic_mean(self):
        return round(self.amount() / self.len, 3)

    def median(self):
        sorted_table = sorted(self.table)
        if self.len % 2 == 0:
            median = (sorted_table[self.len // 2 - 1] + sorted_table[self.len // 2]) * 0.5
        else:
            median = sorted_table[self.len // 2]
        return median

    def fashion(self):
        counts = {}
        for element in self.table:
            if counts.get(element) is None:
                counts[element] = 1
            else:
                counts[element] += 1
        fashion = max(counts.items(), key=lambda x: x[1])[0]
        return fashion

    def min(self):
        return min(self.table)

    def max(self):
        return max(self.table)

    def spread(self):
        return self.max() - self.min()

    def dispersion(self):
        arithmetic_mean = self.arithmetic_mean()
        deviations = list(map(lambda x: x - arithmetic_mean, self.table))
        squares_of_deviations = list(map(lambda x: x ** 2, deviations))
        dispersion = round(sum(squares_of_deviations) / self.len, 3)
        return dispersion

    def standard_deviation(self):
        return round(math.sqrt(self.dispersion()), 3)
